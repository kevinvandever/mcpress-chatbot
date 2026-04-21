"""
Excel Lookup Service for Article Metadata Quality
Feature: article-metadata-quality

Loads metadata from Excel spreadsheets into memory and provides
O(1) lookup by numeric article ID. Supports dual-spreadsheet loading:
- export_subset_DMU_v2.xlsx (14K+ article records)
- MC Press Books - URL-Title-Author.xlsx (115 book records)

Used by MetadataResolver to resolve article titles and authors
from numeric filenames during ingestion and backfill.
"""

import os
import re
import logging
from dataclasses import dataclass, field
from typing import Dict, List, Optional

import openpyxl

logger = logging.getLogger(__name__)


@dataclass
class ExcelMetadataEntry:
    """Represents a metadata entry loaded from an Excel spreadsheet."""
    title: str
    author: str  # raw author string, may contain multiple authors
    url: Optional[str] = None
    article_id: Optional[str] = None


class ExcelLookupService:
    """
    Loads and provides lookup for article metadata from Excel spreadsheets.

    Builds an in-memory dict keyed by numeric article ID string for O(1) lookups.
    Handles two spreadsheet sources:
    1. export_subset_DMU_v2.xlsx — primary article metadata (ID, title, author, URL)
    2. MC Press Books - URL-Title-Author.xlsx — supplementary book metadata
    """

    def __init__(
        self,
        article_spreadsheet_path: str = "export_subset_DMU_v2.xlsx",
        book_spreadsheet_path: str = "MC Press Books - URL-Title-Author.xlsx",
    ):
        """
        Load spreadsheet data into memory on initialization.

        Args:
            article_spreadsheet_path: Path to the article export spreadsheet.
            book_spreadsheet_path: Path to the book metadata spreadsheet.
        """
        self._article_mapping: Dict[str, ExcelMetadataEntry] = {}
        self._book_entries: List[ExcelMetadataEntry] = []

        self._load_article_spreadsheet(article_spreadsheet_path)
        self._load_book_spreadsheet(book_spreadsheet_path)

        logger.info(
            "ExcelLookupService initialized: %d article entries, %d book entries",
            len(self._article_mapping),
            len(self._book_entries),
        )

    # ------------------------------------------------------------------
    # Public lookup methods
    # ------------------------------------------------------------------

    def lookup_by_id(self, article_id: str) -> Optional[ExcelMetadataEntry]:
        """
        Look up metadata by numeric article ID.

        Args:
            article_id: Numeric article ID string (e.g. "27814").

        Returns:
            ExcelMetadataEntry if found, None otherwise.
        """
        return self._article_mapping.get(article_id)

    def lookup_by_filename(self, filename: str) -> Optional[ExcelMetadataEntry]:
        """
        Extract numeric ID from filename and look up metadata.

        Args:
            filename: PDF filename such as "27814.pdf".

        Returns:
            ExcelMetadataEntry if a match is found, None otherwise.
        """
        extracted_id = self.extract_id_from_filename(filename)
        if extracted_id is None:
            return None
        return self.lookup_by_id(extracted_id)

    # ------------------------------------------------------------------
    # Author parsing
    # ------------------------------------------------------------------

    @staticmethod
    def parse_authors(author_string: str) -> List[str]:
        """
        Parse comma / "and" / semicolon-separated author strings into
        individual author names.

        Reuses the splitting pattern from ExcelImportService.parse_authors().

        Args:
            author_string: Raw author string, e.g. "Alice Smith, Bob Jones and Carol Lee".

        Returns:
            List of trimmed individual author names.
        """
        if not author_string or not author_string.strip():
            return []

        s = author_string

        # Collapse "and and" into single "and"
        s = re.sub(r'\s+and\s+and\s+', ' and ', s, flags=re.IGNORECASE)

        # Replace " and " with comma for uniform splitting
        s = re.sub(r'\s+and\s+', ',', s, flags=re.IGNORECASE)

        # Replace semicolons with commas
        s = s.replace(';', ',')

        # Collapse multiple consecutive commas
        s = re.sub(r',\s*,+', ',', s)

        # Split and clean
        authors = [a.strip() for a in s.split(',')]
        authors = [a for a in authors if a]

        return authors

    # ------------------------------------------------------------------
    # ID extraction helpers
    # ------------------------------------------------------------------

    @staticmethod
    def extract_id_from_url(url: str) -> Optional[str]:
        """
        Extract a numeric article ID from a URL path segment.

        Looks for a purely numeric segment in the URL path. For example:
            https://www.mcpressonline.com/programming/rpg/27814-some-slug
            → "27814"

        Also handles trailing numeric segments like:
            https://www.mcpressonline.com/.../12345
            → "12345"

        Args:
            url: Full URL string.

        Returns:
            Numeric ID string if found, None otherwise.
        """
        if not url or not url.strip():
            return None

        # Try to find a numeric ID at the start of a path segment
        # Pattern: digits followed by optional dash+slug, at a path boundary
        match = re.search(r'/(\d+)(?:-[^/]*)?(?:/|$|\?|#)', url)
        if match:
            return match.group(1)

        # Fallback: purely numeric final path segment
        match = re.search(r'/(\d+)\s*$', url)
        if match:
            return match.group(1)

        return None

    @staticmethod
    def extract_id_from_filename(filename: str) -> Optional[str]:
        """
        Extract a numeric ID from a filename like "27814.pdf".

        Args:
            filename: Filename string (e.g. "27814.pdf").

        Returns:
            Numeric ID string if the filename starts with digits, None otherwise.
        """
        if not filename or not filename.strip():
            return None

        # Match leading digits (the numeric article ID)
        match = re.match(r'^(\d+)', filename.strip())
        if match:
            return match.group(1)

        return None

    # ------------------------------------------------------------------
    # Spreadsheet loading (private)
    # ------------------------------------------------------------------

    def _load_article_spreadsheet(self, path: str) -> None:
        """
        Read export_subset_DMU_v2.xlsx using openpyxl and build the
        article ID → ExcelMetadataEntry mapping.

        Column layout:
            A (index 0) = article ID
            B (index 1) = title
            J (index 9) = author
            K (index 10) = article URL
        """
        if not path or not os.path.isfile(path):
            logger.warning(
                "Article spreadsheet not found at '%s' — continuing without article lookup.",
                path,
            )
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active  # Use the active (first) sheet

            rows_loaded = 0
            for row in ws.iter_rows(min_row=2):  # skip header
                # Safely read cells — row may be shorter than expected
                raw_id = row[0].value if len(row) > 0 else None
                raw_title = row[1].value if len(row) > 1 else None
                raw_author = row[9].value if len(row) > 9 else None
                raw_url = row[10].value if len(row) > 10 else None

                if raw_id is None:
                    continue

                article_id = str(raw_id).strip()
                if not article_id:
                    continue

                title = str(raw_title).strip() if raw_title else ""
                author = str(raw_author).strip() if raw_author else ""
                url = str(raw_url).strip() if raw_url else None

                # Skip entries with nan-like values
                if title.lower() == "nan":
                    title = ""
                if author.lower() == "nan":
                    author = ""
                if url and url.lower() == "nan":
                    url = None

                self._article_mapping[article_id] = ExcelMetadataEntry(
                    title=title,
                    author=author,
                    url=url,
                    article_id=article_id,
                )
                rows_loaded += 1

            wb.close()
            logger.info(
                "Loaded %d article entries from '%s'.",
                rows_loaded,
                path,
            )

        except Exception as e:
            logger.warning(
                "Failed to read article spreadsheet '%s': %s — continuing without article lookup.",
                path,
                e,
            )

    def _load_book_spreadsheet(self, path: str) -> None:
        """
        Read MC Press Books - URL-Title-Author.xlsx using openpyxl and
        store as supplementary book entries.

        Expected columns (by header name): URL, Title, Author.
        """
        if not path or not os.path.isfile(path):
            logger.warning(
                "Book spreadsheet not found at '%s' — continuing without book lookup.",
                path,
            )
            return

        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
            ws = wb.active

            # Read header row to find column indices
            header_row = next(ws.iter_rows(min_row=1, max_row=1))
            headers = {
                str(cell.value).strip().lower(): idx
                for idx, cell in enumerate(header_row)
                if cell.value is not None
            }

            url_idx = headers.get("url")
            title_idx = headers.get("title")
            author_idx = headers.get("author")

            if title_idx is None or author_idx is None:
                logger.warning(
                    "Book spreadsheet '%s' missing required columns (Title, Author). "
                    "Found headers: %s",
                    path,
                    list(headers.keys()),
                )
                wb.close()
                return

            rows_loaded = 0
            for row in ws.iter_rows(min_row=2):
                raw_title = row[title_idx].value if len(row) > title_idx else None
                raw_author = row[author_idx].value if len(row) > author_idx else None
                raw_url = row[url_idx].value if url_idx is not None and len(row) > url_idx else None

                title = str(raw_title).strip() if raw_title else ""
                author = str(raw_author).strip() if raw_author else ""
                url = str(raw_url).strip() if raw_url else None

                if title.lower() == "nan":
                    title = ""
                if author.lower() == "nan":
                    author = ""
                if url and url.lower() == "nan":
                    url = None

                if not title and not author:
                    continue

                entry = ExcelMetadataEntry(
                    title=title,
                    author=author,
                    url=url,
                )

                self._book_entries.append(entry)

                # If the URL contains a numeric ID, also index it in the article mapping
                # (only if not already present — article spreadsheet takes precedence)
                if url:
                    extracted_id = self.extract_id_from_url(url)
                    if extracted_id and extracted_id not in self._article_mapping:
                        entry_with_id = ExcelMetadataEntry(
                            title=title,
                            author=author,
                            url=url,
                            article_id=extracted_id,
                        )
                        self._article_mapping[extracted_id] = entry_with_id

                rows_loaded += 1

            wb.close()
            logger.info(
                "Loaded %d book entries from '%s'.",
                rows_loaded,
                path,
            )

        except Exception as e:
            logger.warning(
                "Failed to read book spreadsheet '%s': %s — continuing without book lookup.",
                path,
                e,
            )

    # ------------------------------------------------------------------
    # Introspection helpers (useful for diagnostics)
    # ------------------------------------------------------------------

    @property
    def article_count(self) -> int:
        """Number of article entries loaded."""
        return len(self._article_mapping)

    @property
    def book_count(self) -> int:
        """Number of book entries loaded."""
        return len(self._book_entries)

    @property
    def book_entries(self) -> List[ExcelMetadataEntry]:
        """Read-only access to loaded book entries."""
        return list(self._book_entries)
